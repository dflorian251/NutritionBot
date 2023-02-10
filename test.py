import discord
import bot
from openpyxl import Workbook, load_workbook

def LinearSearch(x):
    foods = []
    max = 0
    for j in range(2, 1617):
        if x.upper() in ws['C'+str(j)].value.upper()  and (max<3):
            foods.append(j)
            max = max+1
    if(len(foods)):
        return foods

    return 1618

def BinarySearch(x,low,high):
    foods = []
    max = 0
    if high >= low and max<3:
 
        mid = (high + low) // 2
 
        # If element is present at the middle itself
        if x.upper() in ws['C'+str(mid)].value.upper():
            foods.append(mid)
            max = max+1
 
        # If element is smaller than mid, then it can only
        # be present in left subarray
        elif ws['C'+str(mid)].value.upper() > x.upper():
            BinarySearch(x,low,mid)
 
        # Else the element can only be present in right subarray
        else:
            BinarySearch(x,mid+1,high)
 
    else:
        return foods

intents = discord.Intents.all()
client = discord.Client(command_prefix='!', intents=intents)
wb = load_workbook("Release 2 - Nutrient file.xlsx")
ws = wb.active

@client.event
async def on_ready():
  print('We have logged in as {0.user}'.format(client))


@client.event
async def on_message(message):
    if message.author == client.user:
        return

    if message.content.startswith("/howto") : 
        await message.channel.send('Για να λάβεις την διατροφική αξία ενός φαγητού γράψε:\n !search !όνομα_φαγητού !γραμμάρια')

    if message.content.startswith("/search"):
        x = str(message.content)
        f = x.split('/',3)
        gram = float(f[3])
        x = str(f[2])
        foods= LinearSearch(x)
        if(foods==1618):
            print("Not found")
        else:
            await message.channel.send("Αποτελέσματα:")
            for food in foods:
                kcal = ( ws["D"+str(food)].value / 4.184 ) * gram /100
                carbs = (ws["AM"+str(food)].value * gram ) / 100
                protein = ( ws["G"+str(food)].value * gram ) /100
                fats = ( ws["I"+str(food)].value * gram ) / 100
                await message.channel.send("%s" %ws["C"+str(food)].value + " nutrition facts per %.1f" %gram + "g\nCalories: %.2f" %kcal + "kcal\nCarbohydrates: %.2f"%carbs + "g\nProtein: %.2f" %protein + "g\nFats: %.2f" %fats + "g\n---------------")


TOKEN = "MTA3MzI1ODIyMDI2NjcyOTU1Mg.GI86Pn.y96EQL-UeJqGgs_2IjoZkbgpmaIawjKNXx1LUU"
client.run(TOKEN)